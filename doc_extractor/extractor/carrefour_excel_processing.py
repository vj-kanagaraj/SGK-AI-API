import pandas as pd
import numpy as np
from functools import reduce
from xlsxwriter.utility import xl_col_to_name
from langid import classify
from sklearn.neural_network import MLPClassifier
import time
from laserembeddings import Laser
import openpyxl
from openpyxl import load_workbook
import pickle
from textblob import TextBlob
import joblib
import warnings
import smbclient
import io

from .excel_processing import *

warnings.filterwarnings("ignore")

from environment import MODE

if MODE == 'local':
    from .local_constants import *
else:
    from .dev_constants import *

laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)

output_file = io.BytesIO()

def get_file(file):
    if file.startswith('\\'):
        print('connecting to SMB share')
        try:
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        except:
            smbclient.reset_connection_cache()
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        finally:
            smbclient.reset_connection_cache()
        return 'SMB'
    else:
        return 'LOCAL'

def excel_extract_carrefour(filepath, sheetname):
    out = get_file(filepath)
    if out == 'SMB':
        wb = openpyxl.load_workbook(output_file)
    else:
        file = document_location + filepath
        wb = openpyxl.load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    t5 = time.time()
    # wb = load_workbook(filepath)
    sheets = wb.sheetnames
    if sheetname in sheets:
        #     sheet= sheets
        ws = wb[sheetname]
        df1 = pd.DataFrame(ws.values)
        df1 = df1.fillna('n/a')
        df1 = df1.apply(lambda x: x.astype(str))
        list1 = df1.values.tolist()
        list1 = [[str(x).replace('N/A', 'n/a').replace('N/a', 'n/a').replace('n/A', 'n/a').replace('NA',
                                                                                                   'n/a').replace(
            '\xa0', ' ').replace('<', '&lt;').replace('>', '&gt;') for x in i] for i in list1]

        list2 = []
        # for i in range(0,len(list1)):
        for j in range(0, len(list1)):
            #         if [list1[i][j]] !='n/a':
            list3 = [list1[j], j + 1]
            list2.append(list3)

        # Split the header and content

        content_list = list2[1:]

        # Keep the language row`

#         header_list = list2[0]

        # Dictionary format

        dic = {}

        for i in range(0, len(content_list)):
            for j in range(0, len(content_list[i][0])):
                if content_list[i][0][j] != 'n/a':
                        cleaned_item = content_list[i][0][j].replace('\n',' ')
                        try:
                            # lang = TextBlob(cleaned_item).detect_language()
                            lang = lang_detect(cleaned_item)
                        except:
                            lang = classify(cleaned_item)[0]
                        # if "#ROW#"+str(content_list[i][1]) !='n/a':
                        if "#ROW#" + str(content_list[i][1]) in dic:
                            dic["#ROW#" + str(content_list[i][1])].append(
                                {xl_col_to_name(j) + str(content_list[i][1]) + '_' + lang: content_list[i][0][j]})
                        else:
                            dic["#ROW#" + str(content_list[i][1])] = [
                                {xl_col_to_name(j) + str(content_list[i][1]) + '_' + lang: content_list[i][0][j]}]

        # classifier model
        classifier = joblib.load(carrefour_model_location)

        # final dictionary
        general_classifier_dic = {}
        for key, value in dic.items():
            #                 if key in cate_classifier:
            for content in value:
                for cell_value, item in content.items():
                    classified_output = classifier.predict(laser.embed_sentences([item], lang='en'))
                    probability1 = classifier.predict_proba(laser.embed_sentences([item], lang='en'))
                    probability1.sort()
                    prob1 = probability1[0][-1]
                    if (prob1 > 0.65) or ((prob1 / 2) > probability1[0][-2]):
                        classified_output = classified_output[0]
                    else:
                        classified_output = 'None'
                    #                             if item not in unwanted_list_items:

                    if classified_output not in ['None', 'others']:
                        if classified_output in general_classifier_dic:
                            general_classifier_dic[classified_output].append({f'{cell_value}': item})
                        else:
                            general_classifier_dic[classified_output] = [{f'{cell_value}': item}]
                    elif classified_output in ['None', 'others']:

                        if key in general_classifier_dic:
                            general_classifier_dic[key].append({f'{cell_value}': item})
                        else:
                            general_classifier_dic[key] = [{f'{cell_value}': item}]

        t6 = time.time()
        print(f'Finished in {t6 - t5} seconds')
        return general_classifier_dic
    else:
        return {'status':'0','comment': 'Invalid Sheetname'}
