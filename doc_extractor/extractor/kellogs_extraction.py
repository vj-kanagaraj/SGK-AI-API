import pandas as pd
import numpy as np
from functools import reduce
from xlsxwriter.utility import xl_col_to_name
from langid import classify
from sklearn.neural_network import MLPClassifier
import time
from laserembeddings import Laser
import openpyxl
from openpyxl import load_workbook
import pickle
from textblob import TextBlob
import joblib
import warnings
import io
import smbclient

warnings.filterwarnings("ignore")

from environment import MODE

if MODE == 'local':
    from .local_constants import *
else:
    from .dev_constants import *

# path_to_bpe_codes = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/93langs.fcodes'
# path_to_bpe_vocab = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/93langs.fvocab'
# path_to_encoder = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/bilstm.93langs.2018-12-26.pt'

laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)

keys1 = ['Brand Name', 'Flavor', 'Marketing Claim', 'Net Contents', 'Product Name', 'Classification', 'Ingredients',
         'Country of Origin', 'Address', 'Best Before Date', 'Contact Information','GDA','Others']

keys2 = ['Energy', 'Protein', 'Total Fat', 'Saturated Fat', 'Carbohydrate', 'Sugars', 'Dietary Fibre', 'Sodium']

keys3 = ['Serving', 'PDV', 'Nutrition Table Content']

cate_classifier = ['Ingredients', 'Best Before Date', 'Address', 'Marketing Claim']

unwanted_list_items = ['Marketing', 'Kellogg', 'CR', 'PS&RA']

# modelname = r'/Users/sakthivel/Documents/SGK/Kelloggs/Python Files/kelloggs_model.sav'

output_file = io.BytesIO()


def get_file(file):
    if file.startswith('\\'):
        print('connecting to SMB share')
        try:
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        except:
            raise Exception('No access for the given path')
        finally:
            smbclient.reset_connection_cache()
        return 'SMB'
    else:
        return 'LOCAL'

def excel_extract_kelloggs(filepath, sheetname):
    print(sheetname)
    input_sheet = sheetname
    t5 = time.time()
    out = get_file(filepath)
    if out == 'SMB':
        wb = openpyxl.load_workbook(output_file)
    else:
        file = document_location + filepath
        wb = openpyxl.load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    # wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        if input_sheet in sheet:
            input_sheet = sheet
    print(input_sheet)
    ws = wb[input_sheet]
    df1 = pd.DataFrame(ws.values)
    df1 = df1.fillna('n/a')
    df1 = df1.apply(lambda x: x.astype(str))
    list1 = df1.values.tolist()
    list1 = [[str(x).replace('N/A', 'n/a').replace('N/a', 'n/a').replace('n/A', 'n/a').replace('NA', 'n/a').replace(
        '\xa0', ' ').replace('<', '&lt;').replace('>', '&gt;') for x in i] for i in list1]

    # Find hidden rows

    hidden_rows = []
    for rowLetter, rowDimension in ws.row_dimensions.items():
        if rowDimension.hidden == True:
            hidden_rows.append(rowLetter)

    # Finding percentage cell format

    cell_index = []
    for row in ws.rows:
        for cell in row:
            if cell.number_format in ['0%', '0.0%', '0.00%']:
                cell_index.append(cell.coordinate)

    # Creating a list with cell/row number

    list2 = []
    # for i in range(0,len(list1)):
    for j in range(0, len(list1)):
        #         if [list1[i][j]] !='n/a':
        list3 = [list1[j], j + 1]
        list2.append(list3)

    # Dropping hidden rows

    list4 = []
    for h in range(0, len(list2)):
        if list2[h][1] not in hidden_rows:
            list4.append(list2[h])

    # Finding an index to pass into classifier

    list5 = []
    # key1 = []
    for i in range(0, len(list4)):
        if list4[i][0][0] != 'n/a':
            val10 = [list4[i][0][0], list4[i][1]]
            list5.append(val10)

    model = joblib.load(kelloggs_model)

    pred_list = []
    for i in range(0, len(list5)):
        prediction = model.predict(laser.embed_sentences(list5[i][0], lang='en'))
        probability = model.predict_proba(laser.embed_sentences(list5[i][0], lang='en'))
        probability.sort()
        prob = probability[0][-1]
        if (prob > 0.65) or ((prob / 2) > probability[0][-2]):
            prediction = prediction[0]
        else:
            prediction = 'None'
        prediction1 = ([list5[i][0], prediction], list5[i][1])
        pred_list.append(prediction1)

    # Creating and splitting list for each category

    nutrition = []
    gen = []
    serving = []

    # Creating a list for each category after prediction

    for i in range(0, len(pred_list)):
        #     for k in range (0,len(list1[i])):
        if pred_list[i][0][1] in keys2:
            nutrition.append(pred_list[i])

        elif pred_list[i][0][1] in keys1:
            gen.append(pred_list[i])

        elif pred_list[i][0][1] in keys3:
            serving.append(pred_list[i])
        else:
            pass

    # New list for nutrition.

    list6 = []
    for i in range(0, len(list2)):
        for k in range(0, len(nutrition)):
            if nutrition[k][1] == list2[i][1]:
                value = list2[i], nutrition[k][0][1]
                list6.append(value)

    # Replacing actual key with Prediction key

    for k in range(0, len(list6)):
        list6[k][0][0][0] = list6[k][1]

    # Getting Keywords for Nutritions

    nutrition_keywords = []
    for y in range(0, len(list6)):
        value1 = [list6[y][0][0][0], list6[y][0][1]]
        nutrition_keywords.append(value1)

    # Creating new list for serving details

    serving_list = []
    for i in range(0, len(list2)):
        for k in range(0, len(serving)):
            if serving[k][1] == list2[i][1]:
                value2 = list2[i], serving[k][0][1]
                serving_list.append(value2)

    # Creating serving dictionary

    serving_dic = {}
    for m in range(0, len(serving_list)):
        for n in range(0, len(serving_list[m][0][0])):
            for key in keys3:
                if key in serving_list[m][1]:
                    if serving_list[m][0][0][n] != 'n/a':
                        lang = classify(serving_list[m][0][0][n])[0]
                        if serving_list[m][1] in serving_dic:
                            serving_dic[serving_list[m][1]].append(
                                {xl_col_to_name(n) + str(serving_list[m][0][1]) + '_' + lang: serving_list[m][0][0][n]})
                        else:
                            serving_dic[serving_list[m][1]] = [
                                {xl_col_to_name(n) + str(serving_list[m][0][1]) + '_' + lang: serving_list[m][0][0][n]}]

    # Nutrition Header
    head = None
    for k in range(0, len(list1)):
        for j in range(0, len(list1[k])):
            if list1[k][j] == '%DI* Per Serving':
                head = list1[k]

    # head = ["PDV" if x == '%DI* Per Serving' else x for x in head]
    heading = []
    for h in head:
        if "%" in h:
            heading.append("PDV")
        else:
            heading.append("Value")

    head = heading

    # Nutrition Dictionary

    nutrition_dic = {}
    for i in range(0, len(list6)):
        for index in keys2:
            if index in list6[i][0][0]:
                index1 = list6[i][0][0].index(index)
                index1 = index1 + 1
                for j in range(index1, len(list6[i][0][0])):
                    if list6[i][0][0][j] != 'n/a':

                        if xl_col_to_name(j) + str(list6[i][0][1]) in cell_index:
                            k = float(list6[i][0][0][j])
                            # k = int(k * 100)
                            k = f"{int(k*100)}%"
                            if list6[i][1] in nutrition_dic:
                                nutrition_dic[list6[i][1]].append(
                                    {head[j]: {xl_col_to_name(j) + str(list6[i][0][1]) + '_en': str(k)}})
                            else:
                                nutrition_dic[list6[i][1]] = [
                                    {head[j]: {xl_col_to_name(j) + str(list6[i][0][1]) + '_en': str(k)}}]
                        else:
                            if list6[i][1] in nutrition_dic:
                                nutrition_dic[list6[i][1]].append(
                                    {head[j]: {xl_col_to_name(j) + str(list6[i][0][1]) + '_en': list6[i][0][0][j]}})
                            else:
                                nutrition_dic[list6[i][1]] = [
                                    {head[j]: {xl_col_to_name(j) + str(list6[i][0][1]) + '_en': list6[i][0][0][j]}}]
                    else:
                        pass
            else:
                pass

    # Merging Nutrition and Serving dictionary

    # new_nutrition_dic = {**nutrition_dic, servin**g_dic}

    # General Category List

    general_category_list = []
    for i in range(0, len(list2)):
        for k in range(0, len(gen)):
            if gen[k][1] == list2[i][1]:
                value4 = list2[i], gen[k][0][1]
                general_category_list.append(value4)
    for k in range(0, len(general_category_list)):
        #             if gen[k][0][1] != 'serving':
        general_category_list[k][0][0][0] = gen[k][0][1]

    # General Category Keywords.

    general_category_keywords = []
    for y in range(0, len(general_category_list)):
        #     for z in range(0,len(nl2[0][0])):
        #         if gen2[y][0][0] !='n/a':
        general_category_keywords.append(general_category_list[y][0][0][0])

    general_category_dic1 = {}
    for i in range(0, len(general_category_list)):
        #     if len(matches2)==len(matches[i]):
        for y in keys1:
            if y in general_category_list[i][0][0]:
                index2 = general_category_list[i][0][0].index(y)
                index2 = index2 + 1
                #                 some = empty[i][0][some+1:]
                #                 nl1.append(some)
                list7 = []
                for k in range(index2, len(general_category_list[i][0][0])):
                    #                     names1= y,[xl_col_to_name(k)+str(empty[i][1]),empty[i][0][k]]
                    if general_category_list[i][0][0][k] != 'n/a':
                        if general_category_list[i][1] in general_category_dic1:
                            general_category_dic1[general_category_list[i][1]].append({xl_col_to_name(k) + str(
                                general_category_list[i][0][1]): general_category_list[i][0][0][k]})
                        else:
                            general_category_dic1[general_category_list[i][1]] = [{xl_col_to_name(k) + str(
                                general_category_list[i][0][1]): general_category_list[i][0][0][k]}]
    #                                 if list2[i][0][k] != 'n/a':
    #                                     value5={xl_col_to_name(k)+str(list2[i][1]):list2[i][0][k]}
    #                                     list7.append(value5)
    #                                     new_dict = {y:list7}
    #                                     general_category_dic.update(new_dict)

    # General Category Dictionary

    general_classifier_dic = {}
    for key, value in general_category_dic1.items():
        if key in cate_classifier:
            for content in value:
                for cell_value, item in content.items():
                    lang = classify(item)[0]
                    #                                 lang = TextBlob(item).detect_language()
                    classified_output = model.predict(laser.embed_sentences([item], lang='en'))
                    probability1 = model.predict_proba(laser.embed_sentences([item], lang='en'))
                    probability1.sort()
                    prob1 = probability1[0][-1]
                    if classified_output == "Ingredients" and prob1 > 0.65:
                        for text in item.split("."):
                            # lang1 = classify(item)[0]
                            text = text.strip()
                            if text:
                                # text = text.strip()
                                classified_output1 = model.predict(laser.embed_sentences([text], lang='en'))
                                if classified_output1[0] in general_classifier_dic:
                                    general_classifier_dic[classified_output1[0]].append({f'{cell_value}_{lang}': text})
                                else:
                                    general_classifier_dic[classified_output1[0]] = [{f'{cell_value}_{lang}': text}]
                        continue

                    if (prob1 > 0.65) or ((prob1 / 2) > probability1[0][-2]):
                        classified_output = classified_output[0]
                    else:
                        classified_output = 'None'
                    if item not in unwanted_list_items:

                        if classified_output not in ['None', 'others']:
                            if classified_output in general_classifier_dic:
                                general_classifier_dic[classified_output].append({f'{cell_value}_{lang}': item})
                            else:
                                general_classifier_dic[classified_output] = [{f'{cell_value}_{lang}': item}]
                        elif classified_output in ['None', 'others']:

                            if key in general_classifier_dic:
                                general_classifier_dic[key].append({f'{cell_value}_{lang}': item})
                            else:
                                general_classifier_dic[key] = [{f'{cell_value}_{lang}': item}]
                    else:
                        pass

        else:
            for content in value:
                for cell_value, item in content.items():
                    lang = classify(item)[0]
                    #                                 lang = TextBlob(item).detect_language()
                    if item not in unwanted_list_items:
                        if key in general_classifier_dic:
                            general_classifier_dic[key].append({f'{cell_value}_{lang}': item})
                        else:
                            general_classifier_dic[key] = [{f'{cell_value}_{lang}': item}]
                    else:
                        pass

    # Attributes Capturing & assigning

    attributes = []
    for row in ws:
        for cell in row:
            if cell.font.bold:
                attributes.append(cell.value)

    #         attributes = [x.lower() for x in attributes]
    attributes = [str(x).replace('\xa0', ' ') for x in attributes]

    final_dic = {}
    for key, value in general_classifier_dic.items():
        for content in value:
            for cell_value, item in content.items():

                if item in attributes:
                    if key in final_dic:
                        final_dic[key].append({f'{cell_value}': '<b>' + item + '</b>'})
                    else:
                        final_dic[key] = [{f'{cell_value}': '<b>' + item + '</b>'}]
                else:
                    if key in final_dic:
                        final_dic[key].append({f'{cell_value}': item})
                    else:
                        final_dic[key] = [{f'{cell_value}': item}]

    final_dic = {**final_dic, **serving_dic}
    final_dic['Nutrition'] = nutrition_dic
    t6 = time.time()
    print(f'Finished in {t6 - t5} seconds')
    return final_dic
