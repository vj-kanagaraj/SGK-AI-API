import pandas as pd
import numpy as np
from functools import reduce
from xlsxwriter.utility import xl_col_to_name
from langid import classify
from sklearn.neural_network import MLPClassifier
import time
from laserembeddings import Laser
import openpyxl
from .excel_processing import *
import io
from textblob import TextBlob

from environment import MODE

if MODE == 'local':
    from .local_constants import *
else:
    from .dev_constants import *

laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)

keys1 = ['Ingredients','Allergen Contains Statement','Allergens May Contain Statement','Handling Statement',
         'Preparation Instructions','Prepared Statement','Storage Type','Product Name','Suffix',
         'Violator','Child Nutrition Statement','Required Copy','Servings Per Container','Serving Size',
         'About (Servings per Container)']

keys2 = ['Calorie','Calorie from Fat','Total Fat','Saturated Fat','Trans Fat','Polyunsaturated Fat','Omega',
         'Monounsaturated Fat','Cholesterol','Sodium','Total Carb','Dietary Fibre','Total Sugar',
         'Inc. Added Sugar','Protein','Vitamin A','Vitamin C','Vitamin D','Calcium','Iron','Potassium',
         'Additional Notes']

cate_classifier = ['Ingredients','Allergen Contains Statement','Allergens May Contain Statement',
                   'Handling Statement','Preparation Instructions','Prepared Statement','Storage Type']

output_file = io.BytesIO()

def get_file(file):
    if file.startswith('\\'):
        print('connecting to SMB share')
        try:
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        except:
            smbclient.reset_connection_cache()
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        finally:
            smbclient.reset_connection_cache()
        return 'SMB'
    else:
        return 'LOCAL'


def excel_extraction_new(filepath, sheetname):
    print('new excel extraction')
    out = get_file(filepath)
    if out == 'SMB':
        wb_obj = openpyxl.load_workbook(output_file)
    else:
        file = document_location + filepath
        wb_obj = openpyxl.load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    if sheetname == "Component Details":
        #         excel_gen_sheet(sheetname)
        # wb_obj = openpyxl.load_workbook(filepath)
        '''ws1 = wb_obj[sheetname]
        df1 = pd.DataFrame(ws1.values)
        df1 = df1.fillna('n/a')
        df1 = df1.apply(lambda x: x.astype(str))

        df2 = df1.transpose()
        newlist = df2.values.tolist()
        newlist = [[str(x).replace('N/A', 'n/a').replace('N/a', 'n/a').replace('n/A', 'n/a').replace('NA',
                                                                                                     'n/a').replace(
            '\xa0', ' ').replace('<', '&lt;').replace('>', '&gt;') for x in i] for i in newlist]

        cellnumber = []
        # for i in range(0,len(list1)):
        for j in range(0, len(newlist)):
            #         if [list1[i][j]] !='n/a':
            list10 = [newlist[j], j]
            cellnumber.append(list10)
        # For Item details
        component_dic = {}
        #         keys5 =['GFS Item #','Brand','GTIN / SCC']
        keys6 = ['Sales Order ID', 'Wave ID', 'Service Order ID']
        for i in range(0, len(cellnumber)):
            #     if len(matches2)==len(matches[i]):
            for y in keys6:
                if y in cellnumber[i][0]:
                    index = cellnumber[i][0].index(y)
                    index = index + 1
                    #                 some = empty[i][0][some+1:]
                    #                 nl1.append(some)
                    list3 = []
                    for k in range(index, len(cellnumber[i][0])):
                        #                     names1= y,[xl_col_to_name(k)+str(empty[i][1]),empty[i][0][k]]
                        if cellnumber[i][0][k] != 'n/a':
                            lang = classify(cellnumber[i][0][k])[0]
                            # lang = TextBlob(cellnumber[i][0][k]).detect_language()
                            values = {f'{xl_col_to_name(cellnumber[i][1]) + str(k + 1)}_{lang}': cellnumber[i][0][k]}

                            #                                 names ={str(cellnumber(k))+xl_col_to_name([i][1]) :cellnumber[i][0][k]}
                            list3.append(values)
                            values1 = {y: list3}
                            component_dic.update(values1)

                else:
                    pass'''
        component_dic = {}
        return component_dic

    elif sheetname == "Item Details":
        #         excel_gen_sheet(sheetname)
        # wb_obj = openpyxl.load_workbook(filepath)
        '''ws1 = wb_obj[sheetname]
        df1 = pd.DataFrame(ws1.values)
        df1 = df1.fillna('n/a')
        df1 = df1.apply(lambda x: x.astype(str))

        df2 = df1.transpose()
        newlist = df2.values.tolist()
        newlist = [[str(x).replace('N/A', 'n/a').replace('N/a', 'n/a').replace('n/A', 'n/a').replace('NA',
                                                                                                     'n/a').replace(
            '\xa0', ' ').replace('<', '&lt;').replace('>', '&gt;') for x in i] for i in newlist]

        cellnumber = []
        # for i in range(0,len(list1)):
        for j in range(0, len(newlist)):
            #         if [list1[i][j]] !='n/a':
            list10 = [newlist[j], j]
            cellnumber.append(list10)
        item_dic = {}
        keys5 = ['GFS Item #', 'Brand', 'GTIN / SCC']
        for i in range(0, len(cellnumber)):
            #     if len(matches2)==len(matches[i]):
            for y in keys5:
                if y in cellnumber[i][0]:
                    index = cellnumber[i][0].index(y)
                    index = index + 1
                    #                 some = empty[i][0][some+1:]
                    #                 nl1.append(some)
                    list4 = []
                    for k in range(index, len(cellnumber[i][0])):
                        #                     names1= y,[xl_col_to_name(k)+str(empty[i][1]),empty[i][0][k]]
                        if cellnumber[i][0][k] != 'n/a':
                            lang = classify(cellnumber[i][0][k])[0]
                            # lang = TextBlob(cellnumber[i][0][k]).detect_language()
                            values2 = {f'{xl_col_to_name(cellnumber[i][1]) + str(k + 1)}_{lang}': cellnumber[i][0][k]}

                            #                                 names ={str(cellnumber(k))+xl_col_to_name([i][1]) :cellnumber[i][0][k]}
                            list4.append(values2)
                            values3 = {y: list4}
                            item_dic.update(values3)

                else:
                    pass'''
        item_dic = {}
        return item_dic

    elif sheetname in wb_obj.sheetnames:

        # wb_obj = openpyxl.load_workbook(filepath)
        ws1 = wb_obj[sheetname]
        df1 = pd.DataFrame(ws1.values)
        df1 = df1.fillna('n/a')
        df1 = df1.apply(lambda x: x.astype(str))
        list2 = df1.values.tolist()
        list2 = [[str(x).replace('N/A', 'n/a').replace('N/a', 'n/a').replace('n/A', 'n/a').replace('NA', 'n/a').replace(
            '\xa0', ' ').replace('<', '&lt;').replace('>', '&gt;') for x in i] for i in list2]

        cell_index = []
        for row in ws1.rows:
            for cell in row:
                if cell.number_format == '0%':
                    cell_index.append(cell.coordinate)

        # Creating a list for content with cell index/number.

        empty = []
        # for i in range(0,len(list1)):
        for j in range(0, len(list2)):
            #         if [list1[i][j]] !='n/a':
            list10 = [list2[j], j + 1]
            empty.append(list10)

        # Get the first valid index in the empty list & append into list5

        list5 = []
        # key1 = []
        for i in range(0, len(empty)):
            #     if empty[i][0] !='n/a':
            #     for j in range(0,len(empty[i][0])):
            #     if empty[i][0][0] !='n/a':
            l = [np.nan if x == 'n/a' else x for x in empty[i][0]]
            nl6 = pd.Series(l)
            nl6 = nl6.first_valid_index()
            if nl6 != None:
                val10 = [empty[i][0][nl6], empty[i][1]]
                list5.append(val10)

        # Loading a pickle(Classifier) file
        # model = pickle.load(open(excel_model_location, 'rb'))
        model = joblib.load(excel_model_location_new)

        #         input_dataset = r"/Users/sakthivel/Documents/SGK/Excel Dataset Samples/New Format/Multi Inputs for AI/Book1.xlsx"
        #          # df = pd.read_excel(input_dataset,sheet_name='main')
        #         df = pd.read_excel(input_dataset,engine='openpyxl')
        #         x_train_laser = laser.embed_sentences(df['text'],lang='en')
        #         classifier = MLPClassifier(hidden_layer_sizes=(80,), solver='adam', activation='tanh', max_iter=750, random_state=0,
        #                      shuffle=True)
        #         classifier.fit(x_train_laser,df['category'])
        #         joblib.dump(classifier,excel_model_location)

        # Predict the labels and append into list1
        #         t9 = time.time()
        list1 = []
        for i in range(0, len(list5)):
            prediction = model.predict(laser.embed_sentences(list5[i][0], lang='en'))
            # probability = model.predict_proba(laser.embed_sentences(list5[i][0], lang='en'))
            # probability.sort()
            # prob = probability[0][-1]
            prediction1 = ([list5[i][0], prediction[0]], list5[i][1])
            list1.append(prediction1)
        #         t10 = time.time()
        #         print(f' completed in {t10-t9} seconds')
        # After predictions, append nutrition and general categories into respective list.

        nutrition = []
        gen = []
        for i in range(0, len(list1)):
            #     for k in range (0,len(list1[i])):
            if list1[i][0][1] in keys2:
                nutrition.append(list1[i])

            elif list1[i][0][1] in keys1:
                gen.append(list1[i])

        nl2 = []
        for i in range(0, len(empty)):
            for k in range(0, len(nutrition)):
                if nutrition[k][1] == empty[i][1]:
                    v = empty[i], nutrition[k][0][1]
                    nl2.append(v)

        # Getting Keywords for Nutritions

        nutrition_keywords = []
        for y in range(0, len(nl2)):
            nutrition_keywords.append(nl2[y][0][0][0])

        # Header to append with nutritions like('Value' & % DV)
        head = None
        for k in range(0, len(list2)):
            for j in range(0, len(list2[k])):
                if list2[k][j].lower().replace(' ', '') == '%dv':
                    head = list2[k]

        head = ["PDV" if x == '% DV' else x for x in head]

        # Nutrtion Dictionary

        #         nutrition_dic={}
        #         for i in range(0,len(nl2)):
        #                 for index in nutrition_keywords:
        #                     if index in nl2[i][0][0]:
        #                         index1 = nl2[i][0][0].index(index)
        #                         index1 = index1+1
        #                         for j in range(index1,len(nl2[i][0][0])):
        #                             if nl2[i][0][0][j] !='n/a':
        #         #                         if matches2[0][j] in k1:
        #                                 if head[j].lower()=='% dv':
        #         #                             if (((len(nl2[i][0][0][j]) == 3  and nl2[i][0][0][j] !='n/a') or len(nl2[i][0][0][j]) == 4) and '.' in nl2[i][0][0][j]):
        #                                     if xl_col_to_name(j)+str(nl2[i][0][1]) in cell_index:
        #         #                             if '.' in nl2[i][0][j]:
        #                                         k = float(nl2[i][0][0][j])
        #                                         k= int(k*100)
        #                                         if index+'_'+ head[j] in nutrition_dic:
        #                     #                         names= {index+'_'+ matches2[0][j]+'_'+str(j):[xl_col_to_name(j)+str(nl2[i][1])+'_en',nl2[i][0][j]]}
        #                                             nutrition_dic[index+'_'+ head[j]].append({xl_col_to_name(j)+str(nl2[i][0][1])+'_en':k})
        #                                         else:
        #                                             nutrition_dic[index+'_'+ head[j]]= [{xl_col_to_name(j)+str(nl2[i][0][1])+'_en':k}]
        #                                     else:

        #                                         if index+'_'+ head[j] in nutrition_dic:
        #                     #                         names= {index+'_'+ matches2[0][j]+'_'+str(j):[xl_col_to_name(j)+str(nl2[i][1])+'_en',nl2[i][0][j]]}
        #                                             nutrition_dic[index+'_'+ head[j]].append({xl_col_to_name(j)+str(nl2[i][0][1])+'_en':nl2[i][0][0][j]})
        #                                         else:
        #                                             nutrition_dic[index+'_'+ head[j]]= [{xl_col_to_name(j)+str(nl2[i][0][1])+'_en':nl2[i][0][0][j]}]
        #                                 else:

        #                                         if index+'_'+ head[j] in nutrition_dic:
        #                     #                         names= {index+'_'+ matches2[0][j]+'_'+str(j):[xl_col_to_name(j)+str(nl2[i][1])+'_en',nl2[i][0][j]]}
        #                                             nutrition_dic[index+'_'+ head[j]].append({xl_col_to_name(j)+str(nl2[i][0][1])+'_en':nl2[i][0][0][j]})
        #                                         else:
        #                                             nutrition_dic[index+'_'+ head[j]]= [{xl_col_to_name(j)+str(nl2[i][0][1])+'_en':nl2[i][0][0][j]}]

        # # Nutrition Final Response

        #         nutrition_final_response ={}
        #         for key, values in nutrition_dic.items():
        #             keywords = key.split('_')[0]
        #             pred = model.predict(laser.embed_sentences(keywords, lang='en'))
        #             if pred[0] in nutrition_final_response:
        #                 nutrition_final_response[pred[0]].append({key:values}) # Use key.split('_')[1] to keep only Value and % DV
        #             else:
        #                 nutrition_final_response[pred[0]]=[{key:values}] # Use key.split('_')[1] to keep only Value and % DV

        # Nutrition Final Response (Dictionary format)
        nutrition_dic = {}
        for i in range(0, len(nl2)):
            for index in nutrition_keywords:
                if index in nl2[i][0][0]:
                    index1 = nl2[i][0][0].index(index)
                    index1 = index1 + 1
                    for j in range(index1, len(nl2[i][0][0])):
                        if nl2[i][0][0][j] != 'n/a':
                            #                         if matches2[0][j] in k1:
                            if head[j].lower() == 'pdv':
                                #                             if (((len(nl2[i][0][0][j]) == 3  and nl2[i][0][0][j] !='n/a') or len(nl2[i][0][0][j]) == 4) and '.' in nl2[i][0][0][j]):
                                if xl_col_to_name(j) + str(nl2[i][0][1]) in cell_index:
                                    #                             if '.' in nl2[i][0][j]:
                                    k = float(nl2[i][0][0][j])
                                    k = int(k * 100)
                                    if nl2[i][1] in nutrition_dic:
                                        #                                             if head[j] in nutrition_dic[nl2[i][1]]:
                                        #                         names= {index+'_'+ matches2[0][j]+'_'+str(j):[xl_col_to_name(j)+str(nl2[i][1])+'_en',nl2[i][0][j]]}
                                        nutrition_dic[nl2[i][1]].append(
                                            {head[j]: {xl_col_to_name(j) + str(nl2[i][0][1]) + '_en': k}})
                                    #                                             else:
                                    #                                                 nutrition_dic[nl2[i][1]][head[j]]= [{xl_col_to_name(j)+str(nl2[i][0][1])+'_en':k}]
                                    else:
                                        nutrition_dic[nl2[i][1]] = [
                                            {head[j]: {xl_col_to_name(j) + str(nl2[i][0][1]) + '_en': k}}]
                                else:
                                    if nl2[i][1] in nutrition_dic:
                                        #                                             if head[j] in nutrition_dic[nl2[i][1]]:
                                        #                         names= {index+'_'+ matches2[0][j]+'_'+str(j):[xl_col_to_name(j)+str(nl2[i][1])+'_en',nl2[i][0][j]]}
                                        nutrition_dic[nl2[i][1]].append(
                                            {head[j]: {xl_col_to_name(j) + str(nl2[i][0][1]) + '_en': nl2[i][0][0][j]}})
                                    #                                             else:
                                    #                                                 nutrition_dic[nl2[i][1]][head[j]]= [{xl_col_to_name(j)+str(nl2[i][0][1])+'_en':nl2[i][0][0][j]}]
                                    else:
                                        nutrition_dic[nl2[i][1]] = [
                                            {head[j]: {xl_col_to_name(j) + str(nl2[i][0][1]) + '_en': nl2[i][0][0][j]}}]
                            else:

                                if nl2[i][1] in nutrition_dic:
                                    #                                             if head[j] in nutrition_dic[nl2[i][1]]:
                                    #                         names= {index+'_'+ matches2[0][j]+'_'+str(j):[xl_col_to_name(j)+str(nl2[i][1])+'_en',nl2[i][0][j]]}
                                    nutrition_dic[nl2[i][1]].append(
                                        {head[j]: {xl_col_to_name(j) + str(nl2[i][0][1]) + '_en': nl2[i][0][0][j]}})
                                #                                             else:
                                #                                                 nutrition_dic[nl2[i][1]][head[j]]= [{xl_col_to_name(j)+str(nl2[i][0][1])+'_en':nl2[i][0][0][j]}]
                                else:
                                    nutrition_dic[nl2[i][1]] = [
                                        {head[j]: {xl_col_to_name(j) + str(nl2[i][0][1]) + '_en': nl2[i][0][0][j]}}]

        # List to get general category from master list and replace a tag with classified output.

        general_category_list = []
        for i in range(0, len(empty)):
            for k in range(0, len(gen)):
                if gen[k][1] == empty[i][1]:
                    v = empty[i], gen[k][0][1]
                    general_category_list.append(v)
        for k in range(0, len(general_category_list)):
            #             if gen[k][0][1] != 'serving':
            general_category_list[k][0][0][0] = gen[k][0][1]

        # General Category Keywords.

        general_category_keywords = []
        for y in range(0, len(general_category_list)):
            #     for z in range(0,len(nl2[0][0])):
            #         if gen2[y][0][0] !='n/a':
            general_category_keywords.append(general_category_list[y][0][0][0])

        # General Category Dictionary

        general_category_dic = {}
        for i in range(0, len(empty)):
            #     if len(matches2)==len(matches[i]):
            for y in general_category_keywords:
                if y in empty[i][0]:
                    index2 = empty[i][0].index(y)
                    index2 = index2 + 1
                    #                 some = empty[i][0][some+1:]
                    #                 nl1.append(some)
                    el = []
                    for k in range(index2, len(empty[i][0])):
                        #                     names1= y,[xl_col_to_name(k)+str(empty[i][1]),empty[i][0][k]]
                        if empty[i][0][k] != 'n/a':
                            names = {xl_col_to_name(k) + str(empty[i][1]): empty[i][0][k]}
                            el.append(names)
                            names1 = {y: el}
                            general_category_dic.update(names1)

        # Feeding general category text into Classifier.

        general_classifier_dic = {}
        for key, value in general_category_dic.items():
            if key in cate_classifier:
                for content in value:
                    for cell_value, item in content.items():
                        lang = classify(item)[0]
                        # lang = TextBlob(item).detect_language()
                        classified_output = model.predict(laser.embed_sentences([item], lang='en'))
                        probability1 = model.predict_proba(laser.embed_sentences([item], lang='en'))
                        probability1.sort()
                        prob1 = probability1[0][-1]
                        if (prob1 > 0.65) or ((prob1 / 2) > probability1[0][-2]):
                            classified_output = classified_output[0]
                        else:
                            classified_output = 'none'

                        if classified_output not in ['none', 'others']:
                            if classified_output in general_classifier_dic:
                                general_classifier_dic[classified_output].append({f'{cell_value}_{lang}': item})
                            else:
                                general_classifier_dic[classified_output] = [{f'{cell_value}_{lang}': item}]
                        elif classified_output in ['none', 'others']:

                            if key in general_classifier_dic:
                                general_classifier_dic[key].append({f'{cell_value}_{lang}': item})
                            else:
                                general_classifier_dic[key] = [{f'{cell_value}_{lang}': item}]
                        else:
                            pass

            else:
                for content in value:
                    for cell_value, item in content.items():
                        lang = classify(item)[0]
                        # lang = TextBlob(item).detect_language()
                        if key in general_classifier_dic:
                            general_classifier_dic[key].append({f'{cell_value}_{lang}': item})
                        else:
                            general_classifier_dic[key] = [{f'{cell_value}_{lang}': item}]

        # Attributes Capturing & assigning

        attributes = []
        for row in ws1:
            for cell in row:
                if cell.font.bold:
                    attributes.append(cell.value)

        #         attributes = [x.lower() for x in attributes]
        attributes = [str(x).replace('\xa0', ' ') for x in attributes]

        final_dic = {}
        for key, value in general_classifier_dic.items():
            for content in value:
                for cell_value, item in content.items():

                    if item in attributes:
                        if key in final_dic:
                            final_dic[key].append({f'{cell_value}': '<b>' + item + '</b>'})
                        else:
                            final_dic[key] = [{f'{cell_value}': '<b>' + item + '</b>'}]
                    else:
                        if key in final_dic:
                            final_dic[key].append({f'{cell_value}': item})
                        else:
                            final_dic[key] = [{f'{cell_value}': item}]

        final_dic['Nutrition'] = nutrition_dic
        # final_dic = {**{'status':1},**final_dic}
        return final_dic
    else:
        return {'status':0,'comment':'page does not exist'}
