from xlsxwriter.utility import xl_col_to_name
import io
import warnings

from .excel_processing import *

from environment import MODE

if MODE == 'local':
    from .local_constants import *
else:
    from .dev_constants import *

warnings.filterwarnings("ignore")

from openpyxl import load_workbook

laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)

output_file = io.BytesIO()

category = ['PLATFORM']
product_isolation = ['Male', 'Female']

def get_file(file):
    if file.startswith('\\'):
        print('connecting to SMB share')
        try:
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        except:
            smbclient.reset_connection_cache()
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        finally:
            smbclient.reset_connection_cache()
        return 'SMB'
    else:
        return 'LOCAL'


def main(filename , sheetname):
    response = {}
    if 'formula' in sheetname.lower().strip():
        response = unilever_doc_e(filename,sheetname)
    elif 'cashable' in sheetname.lower().strip():
        response = unilever_doc_a(filename,sheetname)
    return response

def unilever_doc_a(filename, sheetname):
    out = get_file(filename)
    if out == 'SMB':
        workbook = load_workbook(output_file)
    else:
        file = document_location + filename
        workbook = load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    t5 = time.time()
    sheets = workbook.sheetnames
    if sheetname in sheets:
        worksheet = workbook[sheetname]

        cell_index = []
        for row in worksheet.rows:
            for cell in row:
                cell_content = str(cell.value).replace('<', '&lt;').replace('>', '&gt;')
                if cell.font.bold:
                    temp_list_holder = []
                    for line in cell_content.split('\n'):
                        if line.strip():
                            temp_list_holder.append('<b>' + str(line) + '</b>')
                    cell_index.append({cell.coordinate: '\n'.join(temp_list_holder)})
                else:
                    cell_index.append({cell.coordinate: str(cell_content)})

        # Preprocessing

        df1 = pd.DataFrame(worksheet.values)
        df1 = df1.transpose()
        df1 = df1.fillna('n/a')
        df1 = df1.applymap(str)
        list1 = df1.values.tolist()
        list1 = [[str(x).replace('\xa0', ' ') for x in i] for i in list1]

        # Row Number with respective data

        list2 = []
        for j in range(0, len(list1)):
            list3 = [list1[j], j]
            list2.append(list3)

        header_list = []
        for j in range(0, len(list2)):
            if category[0] in list2[j][0]:
                header_list.append(list2[j][0])

        # General dictionary

        gen_dic = {}
        product_list = []
        for i in range(0, len(list2)):
            for key in product_isolation:
                if key in list2[i][0]:
                    index1 = list2[i][0].index(key)
                    index1 = index1 + 1
                    product_list.append(list2[i][0][index1])
                    # product_list

                    for k in product_list:
                        if k in list2[i][0]:
                            index2 = list2[i][0].index(k)
                            index2 = index2 + 1
                            #                     print(index2)
                            for l in range(index2, len(list2[i][0])):
                                #                         lang = classify(list2[i][0][l])[0]
                                if header_list[0][l] != 'n/a':
                                    if list2[i][0][l] != 'n/a':
                                        if k in gen_dic:
                                            gen_dic[k].append({header_list[0][l]: {
                                                xl_col_to_name(list2[i][1]) + str(l + 1): list2[i][0][l]}})

                                        else:
                                            gen_dic[k] = [{header_list[0][l]: {
                                                xl_col_to_name(list2[i][1]) + str(l + 1): list2[i][0][l]}}]

        # Attribute Capturing

        final_dic = {}
        for k, v in gen_dic.items():
            for value in v:
                for k1, v1 in value.items():
                    for cell_value, content in v1.items():
                        #                 print(content)
                        for c1 in cell_index:
                            for k2, v2 in c1.items():
                                if cell_value in k2:
                                    lang = classify(v2.replace('<b>', '').replace('</b>', '').replace('\n', ''))[0]
                                    if k in final_dic:
                                        final_dic[k].append({k1: {cell_value + '_' + lang: v2}})
                                    else:
                                        final_dic[k] = [{k1: {cell_value + '_' + lang: v2}}]
        t6 = time.time()
        print(f'Finished in {t6 - t5} seconds')
        return final_dic
    else:
        return {'Status': 'Invalid Sheetname'}

def unilever_doc_e(filename, sheetname):
    out = get_file(filename)
    if out == 'SMB':
        workbook = load_workbook(output_file)
    else:
        file = document_location + filename
        workbook = load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    t7 = time.time()
    sheets = workbook.sheetnames
    if sheetname in sheets:
        worksheet = workbook[sheetname]

        cell_index = []
        for row in worksheet.rows:
            for cell in row:
                if cell.font.bold:
                    cell_index.append({cell.coordinate: '<b>' + str(cell.value) + '</b>'})
                else:
                    cell_index.append({cell.coordinate: str(cell.value)})

        # Classifier Model
        #
        # input_dataset = r"/Users/sakthivel/Downloads/dataset (1).xlsx"
        # dataframe = pd.read_excel(input_dataset)
        # x_train_laser = laser.embed_sentences(dataframe['text'], lang='en')
        # classifier = MLPClassifier(hidden_layer_sizes=(80,), solver='adam', activation='tanh', max_iter=750,
        #                            random_state=0,
        #                            shuffle=True)
        # classifier.fit(x_train_laser, dataframe['category'])

        general_classifier_dic = {}

        for dic in cell_index:
            for key, value in dic.items():
                if value not in ['None', '<b>None</b>']:
                    #             lang = TextBlob(item).detect_language()
                    lang = classify(value)[0]
                    output = base('general', model_location).prediction(value)
                    classified_output = output['output']
                    prob1 = output['probability']
                    if prob1 > 0.80:
                        classified_output = classified_output
                    else:
                        classified_output = 'None'
                    #                             if item not in unwanted_list_items:

                    if classified_output in ['ingredients']:
                        if classified_output in general_classifier_dic:
                            general_classifier_dic[classified_output].append({f'{key}_{lang}': value})
                        else:
                            general_classifier_dic[classified_output] = [{f'{key}_{lang}': value}]
                    else:
                        pass
        t8 = time.time()
        print(f'Finished in {t8 - t7} seconds')
        return general_classifier_dic
    else:
        return {'Status': '0','comment': 'sheet name invalid'}
